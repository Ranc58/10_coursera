import os
import random
import requests
from bs4 import BeautifulSoup as bs
from lxml import html
from openpyxl import Workbook, load_workbook


def get_random_courses_list(url):
    courses_quantity = 20
    context = requests.get(url)
    courses_tree = html.fromstring(context.content)
    courses_list = courses_tree.xpath('//loc/text()')
    random_courses_list = random.sample(courses_list, courses_quantity)
    courses_raw_html_list = [requests.get(course).content.decode('utf-8')
                             for course in random_courses_list]
    return courses_raw_html_list


def parse_course_html(courses_raw_html_list):
    parsed_courses_list = []
    for course in courses_raw_html_list:
        parse_info = bs(''.join(course), "lxml")
        course_name = parse_info.find('h1', {'class': 'title display-3-text'})
        start_date = parse_info.find(
            'div', {'class': 'startdate rc-StartDateString caption-text'}
        )
        language = parse_info.find('div', {'class': 'rc-Language'})
        course_duration = len(parse_info.findAll('div', {'class': 'week'}))
        rate = parse_info.find(
            'div', {'class': 'ratings-text bt3-visible-xs'})
        course_info = {
            'course_name': course_name.text,
            'starting_date': start_date.text,
            'course_lang': language.text,
            'course_duration': course_duration,
            'course_rate': rate
        }
        parsed_courses_list.append(course_info)
    return parsed_courses_list


def add_course_content_to_xlsx(xlsx_file, parsed_courses_list):
    courses_info_book = load_workbook(filename=xlsx_file)
    for course in parsed_courses_list:
        if course['course_rate'] is None:
            course['course_rate'] = 'Not rated yet'
        else:
            course['course_rate'] = course['course_rate'].text
        courses_info_sheet = courses_info_book.active
        courses_info_sheet.append([course['course_name'],
                                   course['course_lang'],
                                   course['starting_date'],
                                   course['course_duration'],
                                   course['course_rate']
                                   ])
    return courses_info_book


def output_xlsx(xlsx_file, table_with_course_content):
    table_with_course_content.save(xlsx_file)


def create_xlsx_file_with_template(xlsx_file):
    courses_info_book = Workbook()
    courses_info_sheet = courses_info_book.active
    courses_info_sheet['A1'] = 'Course name'
    courses_info_sheet['B1'] = 'Course language'
    courses_info_sheet['C1'] = 'Start course date'
    courses_info_sheet['D1'] = 'Course duration(on weeks)'
    courses_info_sheet['E1'] = 'Course rate'
    courses_info_book.save(xlsx_file)


if __name__ == '__main__':
    url = 'http://www.coursera.org/sitemap~www~courses.xml'
    xlsx_file = 'coursera.xlsx'
    if not os.path.exists(xlsx_file):
        create_xlsx_file_with_template(xlsx_file)
    print('Please wait.')
    try:
        parsed_courses_list = parse_course_html(get_random_courses_list(url))
    except ValueError:
        print('Error! Please check xml URL')
    except requests.exceptions.ConnectionError:
        print('Error!Please check your connection')
    else:
        table_with_content = add_course_content_to_xlsx(xlsx_file,
                                                        parsed_courses_list)
        output_xlsx(xlsx_file, table_with_content)
        print('Done! Your file "coursera.xlsx" in folder', os.getcwd())
